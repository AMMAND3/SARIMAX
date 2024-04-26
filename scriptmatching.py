import pandas as pd
from openpyxl import load_workbook
import openpyxl
import numpy as np

def find_used_range(sheet):
    # Find the full range of used cells
    min_row = sheet.min_row
    max_row = sheet.max_row
    min_col = sheet.min_column
    max_col = sheet.max_column

    # Convert column numbers to letters
    start_col = openpyxl.utils.get_column_letter(min_col)
    end_col = openpyxl.utils.get_column_letter(max_col)

    # Construct the range string
    data_range = f"{start_col}{min_row}:{end_col}{max_row}"
    return data_range

# Load the workbook and select the active sheet
wb = load_workbook('United 2022 Chunk.xlsx')
sheet = wb.active

# Attempt to find the first table; fall back to used range if no table exists
if sheet.tables:
    tables = sheet.tables
    first_table_name = next(iter(tables))
    first_table = tables[first_table_name]
    data_range = first_table.ref
else:
    # Use the fallback mechanism to find the used range
    data_range = find_used_range(sheet)

# Extract the starting column and row from the data_range
start_cell, end_cell = data_range.replace('$', '').split(':')
start_col = ''.join(filter(str.isalpha, start_cell))
start_row = ''.join(filter(str.isdigit, start_cell))
end_col = ''.join(filter(str.isalpha, end_cell))
header_row_index = int(start_row) - 1
usecols_range = f"{start_col}:{end_col}"

## Now compare tables

# Load the first Excel file. Replace this with your actual file path.
df_first_table = pd.read_excel('United Locations.xlsx')

# Load the second Excel file with the dynamic header and usecols range
df_second_table = pd.read_excel('United 2022 Chunk.xlsx', usecols=usecols_range, header=header_row_index)

# If you find there's a discrepancy, you can strip whitespace from the column names like this:
df_first_table.columns = df_first_table.columns.str.strip()
df_second_table.columns = df_second_table.columns.str.strip()


print(df_first_table.columns.to_list())
print(df_second_table.columns.to_list())


# Ensure the column you want to merge on has the same name in both dataframes
# df_second_table.rename(columns={'Customer Account Number': 'Account Number'}, inplace=True)

# # Merge the first table's Region information into the second table based on Account Number.
# # 'how='left'' ensures all rows from the second table are kept and matched with corresponding rows in the first table
# merged_df = pd.merge(df_second_table, df_first_table[['Account Number', 'Region']], on='Account Number', how='left')

# #print(merged_df.head)

# # Save the merged DataFrame to a new Excel file.
# #merged_df.to_excel('New United Data.xlsx', index=False)


# ### WORKS TILL HERE


# df_prev_rates = pd.read_excel('United Rates Complete.xlsx')

# # Step ii & iii: Match based on region and description
# # Convert Description to string and strip any whitespace for accurate matching
# merged_df['Descritption'] = merged_df['Descritption'].astype(str).str.strip()
# df_prev_rates['Description'] = df_prev_rates['Description'].astype(str).str.strip()

# df_prev_rates.columns = df_prev_rates.columns.str.strip()

# # Step iv: Check 'Date Out Date' lies between the agreement dates
# # Ensure dates are in datetime format
# merged_df['Date Out'] = pd.to_datetime(merged_df['Date Out'])
# df_prev_rates['Agreement Start Date'] = pd.to_datetime(df_prev_rates['Agreement Start Date'])
# df_prev_rates['Agreement End Date'] = pd.to_datetime(df_prev_rates['Agreement End Date'])

# print(df_prev_rates.head)
# desc_counter = 0
# rates_counter = 0

# # Initialize the 'fits' column to 'NO' by default
# merged_df['description_region_year_matches'] = 'NO'
# merged_df['rates_matches'] = 'NO'
# merged_df['rates_diff'] = np.nan

# # Iterate over the merged_df
# for i, row in merged_df.iterrows():
#     # Get rows from df_prev_rates matching region and description
#     mask = (
#         (df_prev_rates['Region'] == row['Region']) &
#         (df_prev_rates['Description'].str.contains(row['Descritption'], case=False, na=False))
#     )
#     matching_rows = df_prev_rates[mask]

#     # Filter rows that match the date criteria
#     date_matches = matching_rows[(matching_rows['Agreement Start Date'] <= row['Date Out']) & (row['Date Out'] <= matching_rows['Agreement End Date'])]

#     if not date_matches.empty:
#         merged_df.at[i, 'description_region_year_matches'] = 'YES'
#         desc_counter += 1

#         # Check which rate is not null and calculate the difference
#         rate_difference_found = False
#         if pd.notnull(row['Day Rate']):
#             rate_difference = date_matches['Daily Rate'] - row['Day Rate']
#             rate_difference_found = True
#         elif pd.notnull(row['Week Rate']):
#             rate_difference = date_matches['Weekly Rate'] - row['Week Rate']
#             rate_difference_found = True
#         elif pd.notnull(row['Month Rate']):
#             rate_difference = date_matches['4 Week Rate'] - row['Month Rate']
#             rate_difference_found = True

#         if rate_difference_found == True:
#             merged_df.at[i, 'rates_diff'] = rate_difference.iloc[0]
#             rate_difference_found = True

#             # Set rates_matches to 'YES' only if the difference is exactly 0
#             if rate_difference.iloc[0] == 0:
#                 merged_df.at[i, 'rates_matches'] = 'YES'
#                 rates_counter += 1

#         # If no valid rate difference was found (i.e., all rates are null), do nothing more
#         if not rate_difference_found:
#             continue


# # Save the updated dataframe back to Excel
# merged_df.to_excel('New United Data.xlsx', index=False)
# print(desc_counter, rates_counter)